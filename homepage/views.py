from django.shortcuts import render
from django.http import HttpResponse
import ssl
from urllib.request import urlopen as uReq
# import requests
from bs4 import BeautifulSoup as soup
import xlrd
import openpyxl
import os
import string
from selenium import webdriver
import time
from .models import Stock, Histr, sStock
import pandas as pd

#########################Calendar##########################
from cal.views import *
from cal.utils import Calendar
from django.utils.safestring import mark_safe
###########################################################

#########################MAIL CHIMP########################
from django.conf import settings
import json
import requests
from .forms import EmailSignupForm
from .models import Signup
from django.contrib import messages
from django.http import HttpResponseRedirect
#########################MAIL CHIMP########################


#########################Authendification##################

#########################Authendification##################


def index(request):
    form = EmailSignupForm()
    context = {
        "home_page": "active", 
        'form': form,
        }

    return render(request, 'homepage/homepage.html', context)

def homepage(request):
    form = EmailSignupForm()

    # use today's date for the calendar

    d = get_date(request.GET.get('day', None))

    # Instantiate ourcalendar class with todays's year and date

    cal = Calendar(d.year, d.month)

    #Call the formatmonth method, which returns our calendar as a table
    html_cal = cal.formatmonth(withyear=True)


    context = {"home_page": "active", 'form': form, "calendar": mark_safe(html_cal),}
    return render(request, 'homepage/homepage.html', context)

def example(request):
    # get_history()
    # store_original()
    # delete_entry("AT0000652011")

    results = Histr.objects.all()
    context = {"ex_page": "active", "results": results}
    return render(request, 'homepage/example.html', context)

def biostock(request):
    # get_history()
    # store_original()
    # delete_entry("AT0000652011")

    results = Histr.objects.all()
    context = {"bio_page": "active", "results": results}
    return render(request, 'homepage/biostock.html', context)

def chemstock(request):
    results = Histr.objects.all()
    context = {"chem_page": "active", "results": results}
    return render(request, 'homepage/chemstock.html', context)

def team(request):
    context = {"team_page": "active"}
    return render(request, 'homepage/team.html', context)

def contact(request):
    context = {"contact_page": "active"}
    return render(request, 'homepage/contact.html', context)

def strategy(request):
    context = {"strategy_page": "active"}
    return render(request, 'homepage/strategy.html', context)

def make_entry(isin, name, price, date):
    s = Histr(isin, name, date, price)
    s.save()
    return

def make_typetwo_entry(ISIN,link,currency,price,markCap,divCap,divCapZero,divCapOne,divCapTwo,divCapThird,divCapFour,divCapFive,divCapSix,resCap,ownZero,ownOne,ownTwo,ownThree,ownFour,ownFive,ownSix,forCap,forZero,forOne,forThree,forFour,forFive,forSix,revOne,revTwo,revThree,revFour,revFive,revSix):
    entry = sStock(ISIN,link,currency,price,markCap,divCap,divCapZero,divCapOne,divCapTwo,divCapThird,divCapFour,divCapFive,divCapSix,resCap,ownZero,ownOne,ownTwo,ownThree,ownFour,ownFive,ownSix,forCap,forZero,forOne,forThree,forFour,forFive,forSix)
    entry.save()
    return

def delete_entry(isin):
    res = Histr.objects.get(isin=isin)
    res.delete()

def store_original():
    xl = pd.ExcelFile('/Users/markmurtagh/todayiinvest/iinvest/homepage/History.xlsx')
    df = xl.parse("Tabelle1")
    for i, row in df.iterrows():
        isin= row[0]
        name= row[1]
        latest= row[2]
        curr = row[3]
        prices = row[4]
        print(isin, name, latest, "current")
        make_entry(isin, name, latest, "current")
        if isinstance(prices, list):
            for j in prices:
                print(str(isin), name, j, "null")
                make_entry(str(isin), name, j, "null")


def get_history():
    input_workbook = xlrd.open_workbook("/Users/markmurtagh/todayiinvest/Barrenberg/Aktie.xlsx")
    input_worksheet = input_workbook.sheet_by_index(0)
    # rawdata_workbook = openpyxl.load_workbook('Hxlsx')
    # rawdata_worksheet = rawdata_workbook.get_sheet_by_name('Tabelle1')
    i = 0
    alphabet = list(string.ascii_uppercase)
    browser = webdriver.Firefox()
    while i <= 630:
        try:
            i = i + 1
            my_url = input_worksheet.cell(i,0).value
            print(my_url)
            finanzen_stockname = my_url.replace("http://www.finanzen.net/aktien/","")
            print('http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie",""))
            guv_url = 'http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie","")

            #url reader function
            uClient = uReq(my_url)
            page_html = uClient.read()
            uClient.close()
            page_soup = soup(page_html, "html.parser")

            #grab isin
            try:
                isin = page_soup.findAll("span",{"class":"instrument-id"})
                isin = isin[0].text
                if isin[0:4] == "ISIN":
                    isin = isin[6:17]
                else:
                    isin = isin[20:32]
                isin_index = 'A' + str(i)
                # rawdata_worksheet[isin_index] = isin
            except Exception as e:
                print('no ISIN')


            #grab name

            try:
                stock_index = 'B' + str(i)
                # rawdata_worksheet[stock_index] = finanzen_stockname.replace("-Aktie","")

            except Exception as e:
                print('no stockname')

    		#grab newest stock price
            try:
                stock = page_soup.findAll("div",{"col-xs-5 col-sm-4 text-sm-right text-nowrap"})
                eur_price = stock[0].contents[0]
                eur_price_index = 'C' + str(i)
                # rawdata_worksheet[eur_price_index] = eur_price
                make_entry(isin, inanzen_stockname.replace("-Aktie",""), eur_price, timezone.now())
            except Exception as e:
                print('no stockdata')

                uClient = uReq(guv_url)
                page_html = uClient.read()
                uClient.close()

                page_soup = soup(page_html, "html.parser")
                guv = page_soup.findAll("td",{"class":"font-bold"})

            try:
                currency = page_soup.findAll("h2",{"class":"box-headline"})
                currency = currency[0].text.split('(in ')
                currency = currency[1]
                currency = currency.replace(')','')
                currency_index = 'D' + str(i)
                # rawdata_worksheet[currency_index] = currency
            except Exception as e:
                print ('no currency')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2009#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child('+ str(x+1) + ') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 4, value = price)
                    make_entry(isin, x+4, datetime(2009, 2, 1), price)
            except Exception as e:
                print('no 2009 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2010#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child('+ str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 16, value = price)
                    make_entry(isin, x+16, datetime(2010, 2, 1), price)
            except Exception as e:
                print('no 2010 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2011#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 28, value = price)
                    make_entry(isin, x+28, datetime(2011, 2, 1), price)
            except Exception as e:
                print('no 2011 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2012#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 40, value = price)
                    make_entry(isin, x+40, datetime(2012, 2, 1), price)
            except Exception as e:
                print('no 2012 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2013#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 52, value = price)
                    make_entry(isin, x+52, datetime(2013, 2, 1), price)
            except Exception as e:
                print('no 2013 data')
            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2014#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 64, value = price)
                    make_entry(isin, x+64, datetime(2014, 2, 1), price)
            except Exception as e:
                print('no 2014 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2015#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 76, value = price)
                    make_entry(isin, x+76, datetime(2015, 2, 1), price)
            except Exception as e:
                print('no 2015 data')
            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2016#jahr')
                for x in range(1,13):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) + ') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 88, value = price)
                    make_entry(isin, x+88, datetime(2016, 2, 1), price)
            except Exception as e:
                print('no 2016 data')

            try:
                browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2017#jahr')
                for x in range(1,9):
                    price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) + ') > td:nth-child(5)').text
                    # rawdata_worksheet.cell(row = i, column = x + 96, value = price)
                    make_entry(isin, x+96, datetime(2017, 2, 1), price)
            except Exception as e:
                print('no 2017 data')
            # rawdata_workbook.save('History.xlsx')
            browser.close()

        except Exception as e:
            print('error')
            print(e)


#########################MAIL CHIMP########################
MAILCHIMP_API_KEY = settings.MAILCHIMP_API_KEY
MAILCHIMP_DATA_CENTER = settings.MAILCHIMP_DATA_CENTER
MAILCHIMP_EMAIL_LIST_ID = settings.MAILCHIMP_EMAIL_LIST_ID

api_url = f'https://{MAILCHIMP_DATA_CENTER}.api.mailchimp.com/3.0'
members_endpoint = f'{api_url}/lists/{MAILCHIMP_EMAIL_LIST_ID}/members'

def subscribe(email):
    data ={
        "email_adress": email,
        "status": "subscribed"
    }

    r = requests.post(
        members_endpoint,
        auth = ("", MAILCHIMP_API_KEY),
        data = json.dumps(data)
    )
    return r.status_code, r.json()

def email_list_signup(request):
    form = EmailSignupForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            email_signup_qs = Signup.objects.filter(email=form.instance.email)
            if email_signup_qs.exists():
                messages.info(request, "You are already subscribed")
            else:
                subscribe(form.instance.email)
                form.save()
    return HttpResponseRedirect(request.META.get("HTTP_REFERER"))